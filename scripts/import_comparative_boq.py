#!/usr/bin/env python3
"""
OpenConstructionERP — comparative BOQ converter.

Reads a "сравнительная ведомость" (multi-variant comparison sheet: name/name_ru/unit
followed by repeating [qty, rate, cost] blocks per variant), extracts ONE variant,
converts it to the normalized import format and posts it to the core-py API.

Usage:
  python3 scripts/import_comparative_boq.py FILE.xlsx --variant 3 \
      --project-code TTZ-V3 --project-name "Pushkin-TTZ Variant 3" \
      [--api http://localhost:8000] [--sheet 0] [--dry-run]

Format assumptions (auto-detected):
  * header row contains variant labels ("Вариант N" / "Variant N"), the next row
    contains Кол-во / Ед.расценка / Стоимость triplets;
  * section rows: short code in col A (e.g. "1А"), no unit;
  * item rows: unit present; qty/rate may be '-' — then derived from cost.
"""
import argparse
import io
import json
import re
import sys
import urllib.request

import openpyxl

NUM = (int, float)


def num(v):
    if isinstance(v, NUM):
        return float(v)
    if v is None:
        return None
    s = str(v).replace("\u00a0", "").replace(" ", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None


def find_layout(rows):
    """Locate the variant header row and the qty/rate/cost column triplets."""
    for i, r in enumerate(rows[:15]):
        labels = [(j, str(c)) for j, c in enumerate(r) if c and re.search(
            r"вариант|variant", str(c), re.I)]
        if labels:
            variants = {}
            for j, lab in labels:
                m = re.search(r"(\d+)", lab)
                if m:
                    variants[int(m.group(1))] = j  # qty col of the triplet
            return i, variants
    raise SystemExit("Variant header row not found")


def extract(path, sheet, variant):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[sheet]] if isinstance(sheet, int) else wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr, variants = find_layout(rows)
    if variant not in variants:
        raise SystemExit(f"Variant {variant} not found; available: {sorted(variants)}")
    qc = variants[variant]           # qty column
    rc, cc = qc + 1, qc + 2          # rate, cost
    items, anomalies, section = [], [], "GEN"
    start = hdr + 2
    # skip a possible numeric ruler row (1 2 3 4 ...)
    if start < len(rows) and str(rows[start][0]).strip() == "1" and str(
            rows[start][2] or "").strip() == "3":
        start += 1
    seq = 0
    for rnum, r in enumerate(rows[start:], start=start + 1):
        a = str(r[0]).strip() if r[0] is not None else ""
        unit = str(r[2]).strip() if r[2] is not None else ""
        if a and not unit and re.fullmatch(r"\d+[\wА-Я]?", a):
            section = a
            continue
        if not unit:
            continue
        name = (r[1] or r[0])
        if not name:
            continue
        qty, rate, cost = num(r[qc]), num(r[rc]), num(r[cc])
        if cost is None and (qty is None or rate is None):
            continue
        if cost is not None and cost > 0:
            if qty is None or qty == 0:
                qty, rate = 1.0, cost
                anomalies.append(f"row {rnum}: qty missing — set qty=1, rate=cost")
            elif rate is None:
                rate = cost / qty
                anomalies.append(f"row {rnum}: rate '-' — derived rate=cost/qty")
            elif abs(qty * rate - cost) > max(1.0, 0.005 * cost):
                anomalies.append(
                    f"row {rnum}: qty*rate={qty*rate:,.2f} != cost={cost:,.2f} "
                    f"(diff {qty*rate-cost:+,.2f}) — kept cost, rate=cost/qty")
                rate = cost / qty
        else:
            if not (qty and rate):
                continue
        seq += 1
        items.append({
            "code": f"{section}-{seq:03d}",
            "name": str(name).strip()[:255],
            "unit": unit[:20] or "ea",
            "quantity": qty,
            "unit_price": round(rate, 6),
            "section": section,
        })
    return items, anomalies


def to_xlsx_bytes(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["code", "name", "unit", "quantity", "unit_price", "section"])
    for it in items:
        ws.append([it["code"], it["name"], it["unit"],
                   it["quantity"], it["unit_price"], it["section"]])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def api_json(url, payload=None, method="GET"):
    req = urllib.request.Request(url, method=method)
    data = None
    if payload is not None:
        req.add_header("Content-Type", "application/json")
        data = json.dumps(payload).encode()
    with urllib.request.urlopen(req, data=data) as resp:
        return json.loads(resp.read())


def api_upload(url, filename, content):
    boundary = "----oceboundary"
    body = (f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; "
            f"filename=\"{filename}\"\r\nContent-Type: application/octet-stream"
            f"\r\n\r\n").encode() + content + f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--variant", type=int, required=True)
    ap.add_argument("--project-code", required=True)
    ap.add_argument("--project-name")
    ap.add_argument("--api", default="http://localhost:8000")
    ap.add_argument("--sheet", type=int, default=0)
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    items, anomalies = extract(a.file, a.sheet, a.variant)
    total = sum(i["quantity"] * i["unit_price"] for i in items)
    print(f"Variant {a.variant}: {len(items)} items, total {total:,.2f}")
    if anomalies:
        print(f"\nData anomalies ({len(anomalies)}):")
        for x in anomalies[:15]:
            print("  !", x)
        if len(anomalies) > 15:
            print(f"  ... and {len(anomalies) - 15} more")
    if a.dry_run:
        return

    projects = api_json(f"{a.api}/api/v1/projects")
    proj = next((p for p in projects if p["code"] == a.project_code), None)
    if not proj:
        proj = api_json(f"{a.api}/api/v1/projects",
                        {"code": a.project_code,
                         "name": a.project_name or a.project_code}, "POST")
        print(f"Created project {proj['code']}")
    res = api_upload(f"{a.api}/api/v1/projects/{proj['id']}/boq/import",
                     "converted.xlsx", to_xlsx_bytes(items))
    print(f"Imported: {res['imported']}, skipped: {res['skipped']}")
    summary = api_json(f"{a.api}/api/v1/projects/{proj['id']}/boq/summary")
    print(f"API total: {summary['base_total']:,.2f}  "
          f"(items in system: {summary['items_count']})")


if __name__ == "__main__":
    main()
