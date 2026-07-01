"""P-01 Executive summary lite: one-call project status workbook for leadership.
Answers the three questions in 30 seconds: where are we, where is the money, what burns."""
import io
import uuid
from datetime import date

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from psycopg.rows import dict_row
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

HDR = PatternFill("solid", fgColor="1F4E79")
HDRF = Font(color="FFFFFF", bold=True)
TITLE = Font(bold=True, size=13)
WARN = PatternFill("solid", fgColor="FFC7CE")


def _pool():
    from app.main import get_pool
    return get_pool()


def style_header(ws, row):
    for c in ws[row]:
        c.fill = HDR; c.font = HDRF; c.alignment = Alignment(horizontal="center")


@router.get("/{project_id}/reports/executive.xlsx")
async def executive_report(project_id: uuid.UUID):
    async with _pool().connection() as conn:
        conn.row_factory = dict_row
        proj = await (await conn.execute(
            "SELECT * FROM projects WHERE id=%s", (project_id,))).fetchone()
        if not proj:
            raise HTTPException(404, "Project not found")

        cost = await (await conn.execute(
            """WITH plan AS (
                 SELECT split_part(c.code,'.',1) ch, sum(i.total_cost) plan
                 FROM boq_items i JOIN cbs_chapters c ON c.id=i.cbs_chapter_id
                 WHERE i.project_id=%(p)s AND i.status<>'cancelled' GROUP BY 1),
               fact AS (
                 SELECT split_part(c.code,'.',1) ch,
                        sum(t.amount_base_currency) FILTER (WHERE t.transaction_type='Actual') actual,
                        sum(t.amount_base_currency) FILTER (WHERE t.transaction_type='Commitment') committed
                 FROM cost_transactions t JOIN cbs_chapters c ON c.id=t.cbs_chapter_id
                 WHERE t.project_id=%(p)s GROUP BY 1)
               SELECT COALESCE(p.ch,f.ch) ch,
                      (SELECT name FROM cbs_chapters WHERE code=COALESCE(p.ch,f.ch)
                        AND project_id IS NULL) nm,
                      COALESCE(p.plan,0) plan, COALESCE(f.committed,0) committed,
                      COALESCE(f.actual,0) actual
               FROM plan p FULL OUTER JOIN fact f ON f.ch=p.ch ORDER BY 1""",
            {"p": project_id})).fetchall()

        earned = await (await conn.execute(
            """SELECT sum(i.total_cost) plan,
                      sum(LEAST(COALESCE(e.done,0)/NULLIF(i.quantity,0),1)*i.total_cost) ev
               FROM boq_items i LEFT JOIN (
                 SELECT boq_item_id, sum(qty_done) done
                 FROM daily_work_entries GROUP BY 1) e ON e.boq_item_id=i.id
               WHERE i.project_id=%s AND i.status<>'cancelled'""",
            (project_id,))).fetchone()

        drives = await (await conn.execute(
            """SELECT d.code, d.name, d.design_rings, d.status,
                      count(r.id) built, max(r.chainage) chainage
               FROM tunnel_drives d LEFT JOIN tunnel_rings r ON r.drive_id=d.id
               WHERE d.project_id=%s GROUP BY d.id ORDER BY d.code""",
            (project_id,))).fetchall()

        rfis = await (await conn.execute(
            """SELECT code, subject, status, due_date, assigned_to,
                      (status='open' AND due_date < CURRENT_DATE) overdue
               FROM rfis WHERE project_id=%s AND status='open'
               ORDER BY due_date NULLS LAST""",
            (project_id,))).fetchall()

        last_dr = await (await conn.execute(
            """SELECT report_date, shift, manpower_total, equipment_total, narrative
               FROM daily_reports WHERE project_id=%s
               ORDER BY report_date DESC, shift LIMIT 5""",
            (project_id,))).fetchall()

    wb = openpyxl.Workbook()

    # -- Sheet 1: Status
    ws = wb.active; ws.title = "Status"
    ws["A1"] = f"{proj['code']} — {proj['name']}"; ws["A1"].font = TITLE
    ws["A2"] = f"Executive status as of {date.today().isoformat()} · currency {proj['currency']}"
    plan = float(earned["plan"] or 0); ev = float(earned["ev"] or 0)
    actual = sum(float(r["actual"]) for r in cost)
    committed = sum(float(r["committed"]) for r in cost)
    kpi = [("Budget (BOQ plan)", plan),
           ("Earned value (physical)", ev),
           ("Physical %", round(100*ev/plan, 2) if plan else 0),
           ("Actual cost to date", actual),
           ("Committed", committed),
           ("Open RFIs", len(rfis)),
           ("Overdue RFIs", sum(1 for r in rfis if r["overdue"]))]
    ws.append([]); ws.append(["KPI", "Value"]); style_header(ws, ws.max_row)
    for k, v in kpi:
        ws.append([k, v])
        if k == "Overdue RFIs" and v:
            ws.cell(ws.max_row, 2).fill = WARN
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 18

    # -- Sheet 2: Cost by chapter
    ws = wb.create_sheet("Cost")
    ws.append(["Ch", "Chapter", "Plan", "Committed", "Actual", "Variance", "Spent %"])
    style_header(ws, 1)
    for r in cost:
        p, a, c = float(r["plan"]), float(r["actual"]), float(r["committed"])
        ws.append([r["ch"], r["nm"], p, c, a, p - a, round(100*a/p, 1) if p else None])
    ws.append(["", "TOTAL",
               f"=SUM(C2:C{ws.max_row})", f"=SUM(D2:D{ws.max_row})",
               f"=SUM(E2:E{ws.max_row})", f"=C{ws.max_row+1}-E{ws.max_row+1}", ""])
    ws.cell(ws.max_row, 2).font = Font(bold=True)
    for col, w in zip("ABCDEFG", (6, 40, 14, 14, 14, 14, 9)):
        ws.column_dimensions[col].width = w

    # -- Sheet 3: Tunnel
    ws = wb.create_sheet("Tunnel")
    ws.append(["Drive", "Name", "Design rings", "Built", "%", "Chainage m", "Status"])
    style_header(ws, 1)
    for d in drives:
        pct = round(100*d["built"]/d["design_rings"], 1) if d["design_rings"] else None
        ws.append([d["code"], d["name"], d["design_rings"], d["built"], pct,
                   float(d["chainage"] or 0), d["status"]])
    for col, w in zip("ABCDEFG", (10, 34, 13, 9, 7, 12, 12)):
        ws.column_dimensions[col].width = w

    # -- Sheet 4: Open RFIs
    ws = wb.create_sheet("Open RFIs")
    ws.append(["Code", "Subject", "Assigned", "Due", "Overdue"])
    style_header(ws, 1)
    for r in rfis:
        ws.append([r["code"], r["subject"], r["assigned_to"],
                   str(r["due_date"] or ""), "YES" if r["overdue"] else ""])
        if r["overdue"]:
            ws.cell(ws.max_row, 5).fill = WARN
    for col, w in zip("ABCDE", (10, 60, 18, 12, 9)):
        ws.column_dimensions[col].width = w

    # -- Sheet 5: Recent daily reports
    ws = wb.create_sheet("Daily")
    ws.append(["Date", "Shift", "Manpower", "Equipment", "Narrative"])
    style_header(ws, 1)
    for d in last_dr:
        ws.append([str(d["report_date"]), d["shift"], d["manpower_total"],
                   d["equipment_total"], (d["narrative"] or "")[:120]])
    for col, w in zip("ABCDE", (12, 8, 10, 10, 70)):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="Executive_{proj["code"]}_{date.today()}.xlsx"'})
