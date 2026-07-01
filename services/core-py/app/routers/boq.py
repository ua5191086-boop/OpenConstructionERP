"""BOQ vertical: list, Excel import (RU/EN headers), summary with regional
coefficients, Excel export. Items are mirrored into the ontology graph."""
import io
import re
import uuid
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from psycopg.rows import dict_row
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.routers.projects import mirror_object, link_objects

router = APIRouter()

# Flexible header mapping: normalised header -> canonical field
HEADER_MAP = {
    "code": "code", "код": "code", "шифр": "code", "поз": "code", "item": "code",
    "name": "name", "наименование": "name", "описание работ": "name",
    "description": "description", "примечание": "description",
    "unit": "unit", "ед": "unit", "едизм": "unit", "uom": "unit",
    "quantity": "quantity", "колво": "quantity", "количество": "quantity", "qty": "quantity",
    "unitprice": "unit_price", "unitrate": "unit_price", "цена": "unit_price",
    "расценка": "unit_price", "rate": "unit_price", "ценазаед": "unit_price",
    "section": "section", "участок": "section",
    "cbs": "cbs", "глава": "cbs", "chapter": "cbs",
    "currency": "currency", "валюта": "currency",
}


def norm_header(h) -> str:
    return re.sub(r"[^a-zа-я]", "", str(h or "").strip().lower())


def _pool():
    from app.main import get_pool
    return get_pool()


@router.get("/{project_id}/boq/items")
async def list_items(
    project_id: uuid.UUID,
    q: Optional[str] = None,
    cbs: Optional[str] = None,
    limit: int = Query(100, le=1000),
    offset: int = 0,
):
    sql = """SELECT i.id, i.code, i.name, i.unit, i.quantity, i.unit_price,
                    i.total_cost, i.currency, i.status,
                    c.code AS cbs_code, c.name AS cbs_name,
                    s.code AS section_code
             FROM boq_items i
             JOIN cbs_chapters c ON c.id = i.cbs_chapter_id
             JOIN boq_objects  bo ON bo.id = i.object_id
             JOIN boq_complexes bc ON bc.id = bo.complex_id
             JOIN boq_sections s ON s.id = bc.section_id
             WHERE i.project_id = %s"""
    args: list = [project_id]
    if q:
        sql += " AND (i.name ILIKE %s OR i.code ILIKE %s)"
        args += [f"%{q}%", f"%{q}%"]
    if cbs:
        sql += " AND c.code LIKE %s"; args.append(f"{cbs}%")
    sql += " ORDER BY i.code LIMIT %s OFFSET %s"; args += [limit, offset]
    async with _pool().connection() as conn:
        conn.row_factory = dict_row
        return await (await conn.execute(sql, args)).fetchall()


async def _get_or_create_chain(conn, project_id, section_code: str):
    """Ensure section -> complex -> object chain exists; return boq_objects.id."""
    sec = await (await conn.execute(
        """INSERT INTO boq_sections (project_id, code, name, section_type)
           VALUES (%s,%s,%s,'Track')
           ON CONFLICT (project_id, code) DO UPDATE SET updated_at = NOW()
           RETURNING id""",
        (project_id, section_code, section_code),
    )).fetchone()
    cx = await (await conn.execute(
        """INSERT INTO boq_complexes (project_id, section_id, code, name)
           VALUES (%s,%s,%s,%s)
           ON CONFLICT (project_id, code) DO UPDATE SET updated_at = NOW()
           RETURNING id""",
        (project_id, sec["id"], f"{section_code}-CX", f"{section_code} complex"),
    )).fetchone()
    obj = await (await conn.execute(
        """INSERT INTO boq_objects (project_id, complex_id, code, name)
           VALUES (%s,%s,%s,%s)
           ON CONFLICT (project_id, code) DO UPDATE SET updated_at = NOW()
           RETURNING id""",
        (project_id, cx["id"], f"{section_code}-OBJ", f"{section_code} object"),
    )).fetchone()
    return obj["id"]


@router.post("/{project_id}/boq/import")
async def import_boq(project_id: uuid.UUID, file: UploadFile = File(...)):
    """Import BOQ from .xlsx. Recognises RU/EN headers (see HEADER_MAP).
    Required: name, unit, quantity, unit_price. Optional: code, cbs, section."""
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Only .xlsx files are supported")
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        raw_header = next(rows)
    except StopIteration:
        raise HTTPException(400, "Empty file")
    cols = {}
    for idx, h in enumerate(raw_header):
        field = HEADER_MAP.get(norm_header(h))
        if field and field not in cols:
            cols[field] = idx
    missing = {"name", "unit", "quantity", "unit_price"} - cols.keys()
    if missing:
        raise HTTPException(400, f"Missing required columns: {sorted(missing)}. "
                                 f"Recognised headers: {sorted(set(HEADER_MAP.keys()))}")

    imported, skipped, errors = 0, 0, []
    async with _pool().connection() as conn:
        conn.row_factory = dict_row
        proj = await (await conn.execute(
            "SELECT id, code, name FROM projects WHERE id = %s", (project_id,)
        )).fetchone()
        if not proj:
            raise HTTPException(404, "Project not found")
        proj_obj = await mirror_object(conn, "project", project_id,
                                       {"code": proj["code"], "name": proj["name"]},
                                       "projects", project_id)
        cbs_cache, chain_cache = {}, {}

        for rnum, r in enumerate(rows, start=2):
            def cell(f, default=None):
                i = cols.get(f)
                return r[i] if i is not None and i < len(r) else default
            name = cell("name")
            if name is None or str(name).strip() == "":
                skipped += 1
                continue
            try:
                qty = float(cell("quantity") or 0)
                price = float(cell("unit_price") or 0)
            except (TypeError, ValueError):
                errors.append(f"row {rnum}: quantity/price not numeric"); skipped += 1
                continue
            code = str(cell("code") or f"IMP-{rnum:05d}").strip()
            unit = str(cell("unit") or "ea").strip()[:20]
            currency = str(cell("currency") or "USD").strip()[:3].upper()
            section = str(cell("section") or "GEN").strip()[:50]
            cbs_code = str(cell("cbs") or "12.07").strip()

            if cbs_code not in cbs_cache:
                ch = await (await conn.execute(
                    """SELECT id FROM cbs_chapters
                       WHERE code = %s AND project_id IS NULL LIMIT 1""",
                    (cbs_code,),
                )).fetchone()
                if not ch:
                    ch = await (await conn.execute(
                        """SELECT id FROM cbs_chapters
                           WHERE code = '12.07' AND project_id IS NULL""",
                    )).fetchone()
                cbs_cache[cbs_code] = ch["id"]
            if section not in chain_cache:
                chain_cache[section] = await _get_or_create_chain(conn, project_id, section)

            item = await (await conn.execute(
                """INSERT INTO boq_items (project_id, object_id, cbs_chapter_id, code,
                                          name, unit, quantity, unit_price, currency)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   RETURNING id, total_cost""",
                (project_id, chain_cache[section], cbs_cache[cbs_code], code,
                 str(name).strip()[:255], unit, qty, price, currency),
            )).fetchone()
            item_obj = await mirror_object(
                conn, "boq_item", project_id,
                {"code": code, "name": str(name)[:120], "total": float(item["total_cost"] or 0)},
                "boq_items", item["id"])
            await link_objects(conn, "belongs_to", item_obj, proj_obj)
            imported += 1

    return {"imported": imported, "skipped": skipped,
            "errors": errors[:20], "project": proj["code"]}


@router.get("/{project_id}/boq/summary")
async def boq_summary(project_id: uuid.UUID, region: str = "BASE"):
    async with _pool().connection() as conn:
        conn.row_factory = dict_row
        rc = await (await conn.execute(
            "SELECT * FROM regional_coefficients WHERE region_code = %s", (region.upper(),)
        )).fetchone()
        if not rc:
            raise HTTPException(400, f"Unknown region: {region}")
        chapters = await (await conn.execute(
            """SELECT split_part(c.code, '.', 1) AS chapter,
                      min(pc.name) AS chapter_name,
                      count(*) AS items,
                      sum(i.total_cost) AS total
               FROM boq_items i
               JOIN cbs_chapters c ON c.id = i.cbs_chapter_id
               LEFT JOIN cbs_chapters pc
                 ON pc.code = split_part(c.code, '.', 1) AND pc.project_id IS NULL
               WHERE i.project_id = %s AND i.status <> 'cancelled'
               GROUP BY 1 ORDER BY 1""",
            (project_id,),
        )).fetchall()
        base_total = sum(float(ch["total"] or 0) for ch in chapters)
        factor = float(rc["overall_factor"])
        for ch in chapters:
            ch["total"] = float(ch["total"] or 0)
            ch["total_adjusted"] = round(ch["total"] * factor, 2)
    return {
        "region": rc["region_code"], "region_name": rc["region_name"],
        "overall_factor": factor,
        "items_count": sum(ch["items"] for ch in chapters),
        "base_total": round(base_total, 2),
        "adjusted_total": round(base_total * factor, 2),
        "by_chapter": chapters,
    }


@router.get("/{project_id}/boq/export.xlsx")
async def export_boq(project_id: uuid.UUID, region: str = "BASE"):
    async with _pool().connection() as conn:
        conn.row_factory = dict_row
        proj = await (await conn.execute(
            "SELECT code, name, currency FROM projects WHERE id = %s", (project_id,)
        )).fetchone()
        if not proj:
            raise HTTPException(404, "Project not found")
        rc = await (await conn.execute(
            "SELECT * FROM regional_coefficients WHERE region_code = %s", (region.upper(),)
        )).fetchone() or {"overall_factor": 1, "region_code": "BASE"}
        items = await (await conn.execute(
            """SELECT i.code, i.name, i.unit, i.quantity, i.unit_price, i.total_cost,
                      c.code AS cbs
               FROM boq_items i JOIN cbs_chapters c ON c.id = i.cbs_chapter_id
               WHERE i.project_id = %s AND i.status <> 'cancelled'
               ORDER BY c.code, i.code""",
            (project_id,),
        )).fetchall()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BOQ"
    hdr_fill = PatternFill("solid", fgColor="1F4E79"); hdr_font = Font(color="FFFFFF", bold=True)
    ws.append([f"BOQ — {proj['code']} {proj['name']}  |  region: {rc['region_code']} "
               f"(factor {float(rc['overall_factor']):.3f})  |  currency: {proj['currency']}"])
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["CBS", "Code", "Name", "Unit", "Quantity", "Unit Price", "Total", "Total (adj.)"]
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    f = float(rc["overall_factor"]); total = adj = 0.0
    for it in items:
        t = float(it["total_cost"] or 0); total += t; adj += t * f
        ws.append([it["cbs"], it["code"], it["name"], it["unit"],
                   float(it["quantity"]), float(it["unit_price"]), t, round(t * f, 2)])
    ws.append([]); ws.append(["", "", "TOTAL", "", "", "", round(total, 2), round(adj, 2)])
    ws[ws.max_row][2].font = Font(bold=True)
    for col, w in zip("ABCDEFGH", (10, 14, 60, 8, 12, 14, 16, 16)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="BOQ_{proj["code"]}_{rc["region_code"]}.xlsx"'},
    )
